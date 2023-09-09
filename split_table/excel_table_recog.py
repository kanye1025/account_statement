import json
import os
from tools.embeding_tool_table import EmbedingToolTable as EmbedingTool
from config.config import CONF
from langchain.embeddings.huggingface import HuggingFaceEmbeddings
from openpyxl import load_workbook

import  xlrd
from copy import deepcopy
import datetime
from io import BytesIO
import numpy as np
class ExcelTableRecog:
    heads = "流水号/订单号,时间,收入,支出,备注/用途,余额/金额,对方账号,对方户名,收付款方式,开户机构/开户行,分类"
    #embeding = HuggingFaceEmbeddings(model_name=CONF.embeding_model_path)
    #h_emb = embeding.embed_query(heads)
    #h_emb = None
    #embeding = None
    def __init__(self,file_path,byte = None):
        EmbedingTool.init()
        _,ext = os.path.splitext(file_path)
        formatting_xlrd =  ext == '.xls'

        
        if byte:
            self.book = xlrd.open_workbook(file_contents=byte, formatting_info=formatting_xlrd)
            self.booksheet = self.book.sheets()[0]
            if not formatting_xlrd:
                file_path = BytesIO(byte)
        else:
            self.book = xlrd.open_workbook(file_path, formatting_info=formatting_xlrd)
            self.booksheet = self.book.sheets()[0]
     
        self.ws = load_workbook(file_path, read_only=True).active  if not formatting_xlrd else None
        self.nrows = self.booksheet.nrows
        self.ncols = self.booksheet.ncols

        
            
            
        self.data = list()
        for i in range(self.booksheet.nrows):
            col = list()
            for j in range(self.booksheet.ncols):
                value = self.booksheet.cell_value(i,j)
 
                col.append(self.clear_value(value))
            self.data.append(col)
        '''
        for col in datas:
            line = list()
            for v in col:
                if v:
                    line.append(v)
            line = ','.join(line)
            print(line)
        '''

    def is_cell_datetime(self,i, j):
        if self.ws:
            return self.ws.cell(row=i + 1, column=j + 1).is_date
    
        else:
            format = self.book.xf_list[self.booksheet.cell(i, j).xf_index]
            format = self.book.format_map[format.format_key]
            return format.type == 1
    def process_cell_datatime(self, value):
        
        date = datetime.datetime(year=1899, month=12, day=30) + datetime.timedelta(days=float(value))
        int_part,deci_part = str(value).split('.')
        int_part = int(int_part)
        deci_part = int(deci_part)
        if int_part and not deci_part:
            return date.strftime("%Y/%m/%d")
        if not int_part and deci_part:
            return date.strftime("%H:%M:%S")
        if  int_part and deci_part:
            return date.strftime("%Y/%m/%d %H:%M:%S")
        
    def clear_value(self,value):
        return str(value).strip().strip('\t')
    
    
    def get_col_index_range(self,head_index):
        col_index_range = [200, -1]
        for j in range(self.ncols):
            value =self.data[head_index][j]
            if value:
                col_index_range[0] = min(col_index_range[0], j)
                col_index_range[1] = max(col_index_range[1], j + 1)
        for j in range(col_index_range[0],col_index_range[1]):
            if not self.data[head_index][j] :
                raise Exception(f"head col {j} is empty")
        return col_index_range
    def get_row_index_range(self,head_index,col_index_range):
        is_end = True
        for i in range(head_index+1,self.nrows):
            is_table_body = False
            in_col_count = 0
            for j in range(self.ncols):
                value = self.data[i][j]
                if j >=col_index_range[0] and j <col_index_range[1]:
                    if  value:
                        in_col_count+=1
                        if in_col_count>=4:#表格内至少有四列不为空（时间、流水号、金额、交易方）
                            is_table_body = True
                            break
                        continue
                else:
                    if value:#表格外不能有数据
                        is_table_body=False
                        is_end = False
                        break
            if not is_table_body:
                is_end = False
                break
        if is_end:
            row_index_range = [head_index, i+1]
        else:
            row_index_range = [head_index,i]
        return row_index_range
    def get_table_data(self):
        heads = EmbedingTool.get_head_index(self.data)
        datetime_col_indexes = set()
        for col_index in range(self.ncols):
            if self.is_cell_datetime(heads[0]+1,col_index):
                datetime_col_indexes.add(col_index)
        #heads = self.get_head_index()
        #heads = [1, 40, 54, 101, 158, 208, 307, 422, 503, 550, 602, 711]
        #heads = [24]
        col_range = self.get_col_index_range(heads[0])
        row_ranges = [self.get_row_index_range(head, col_range) for head in heads]
      
        row_infos = {}
        for table_index,row_range in enumerate(row_ranges):
            for row in range(row_range[0],row_range[1]):
                row_order = '{:0>3d}'.format(table_index+1)+'_'+str(row-row_range[0]+1)
                header_row =  row in heads
                row_infos[row] = {'row_order':row_order,"header_row":header_row}

        
        ret_obj = {}
        ret_obj["res1"] = {}
        ret_obj["res1"]["outside_infos"] = list()
        ret_obj['res2'] = list()
        for row in range(self.nrows):
            if row in row_infos:#表内
                row_info = row_infos[row]
                row_obj = {}
                row_obj.update(row_info)
                for col in range(col_range[0],col_range[1]):
                #for col,v in enumerate(self.data[row]):
                    v = self.data[row][col]
                    if not row_obj["header_row"] and col in datetime_col_indexes:
                        v = self.process_cell_datatime(v)
                    key = f"{row_info['row_order'].split('_')[1]}.{col-col_range[0]+1}"
                    row_obj[key] = v
                ret_obj['res2'].append(row_obj)
            elif row < heads[0]:                #只取首页表前面的信息
                line =list()
                for value in self.data[row]:
                    if value:
                        line.append(value)
                if line:
                    line = ' '.join(line)
                    ret_obj["res1"]["outside_infos"].append({"txt":line})
                    
        ret_obj['doc_tpye']="excel"
        ret_obj['page_sum'] = len(row_ranges)
        return ret_obj

    
        
        
    def print_table_data(self,row_index_range,col_index_range):
        for i in range(row_index_range[0],row_index_range[1]):
            line = list()
            for j in range(col_index_range[0],col_index_range[1]):
                line.append(self.data[i][j])

            print('\t'.join(line))
            
if __name__ =="__main__":
    #etr = ExcelTableRecog("data/input/test.xls")
    #etr = ExcelTableRecog("data/input/2022攀农业银行1-9月流水.xls")
    #etr = ExcelTableRecog("data/input/甘玉兰化妆品2022.7-2022.9明细.xls")
    etr = ExcelTableRecog("data/input1/攀德中国银行流水2021年.xlsx")
    #etr = ExcelTableRecog("data/alipay_record_20230727_112459.xlsx")
    obj = etr.get_table_data()
    with open('data/out.json','w',encoding='utf-8') as f:
        json_str = json.dump(obj,f,ensure_ascii=False)
    #print(json_str)
    #heads = etr.get_head_index(book_sheet)
    #heads = [1,40,54,101,158,208,307,422,503,550,602,711]
    #print(etr.get_page_seps())
    #col_range = etr.get_col_index_range( heads[0])
    #for head in heads:
    #    print(head[2])
    #for head in heads:
    #    row_range = etr.get_row_index_range(head,col_range)
    #    etr.print_table_data( row_range,col_range)